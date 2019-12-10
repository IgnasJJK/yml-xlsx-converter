from openpyxl import load_workbook, Workbook
import yaml
import sys

def print_usage():
    print('Usage: python3 main.py <input> <output>')
    print('''
    If the input is yml, the output is xlsx in the following format:
    | YML Path | String | Translation |
    
    YML Path (column A) is a collapsed path to the specific string
    in yml.  For example, 'en.homepage.header.homepage'. This is 
    used to convert the xlsx file back into an yml file and must
    not be edited.
    
    String (column B) is the untranslated text.
    
    Translation (column C) is the translated text. When the xlsx 
    file is used as input, this column is used for converting back
    to yml.

    WARNING: There isn't much error-checking.
    ''')

def force_extension(filepath, extension):
    if len(filepath) == 0: 
        raise Exception('Invalid argument: filepath is empty.')
    elif filepath.split('.')[-1].lower() != extension.lower():
        return filepath + '.' + extension
    else:
        return filepath

def crawl_yml_tree(yml, parent=None):
    for heading in yml:
        current_path = '{p}.{head}'.format(p=parent, head=heading) if parent else heading
        value = yml[heading]
        if type(value) is dict:
            for result in crawl_yml_tree(value, parent=current_path):
                yield result
        elif type(value) is list:
            for index, val in enumerate(value):
                yield ('{p}[{ix}]'.format(p=current_path, ix=index), val)
        else:
            yield (current_path, value)


def convert_yml_to_xlsx(yml):

    wb = Workbook()
    ws = wb.active

    ## Write header ##
    ws[PATH_COL + '1'] = 'ymlpath'
    ws[SOURCE_COL + '1'] = 'untranslated'
    ws[TRANSLATED_COL + '1'] = 'translated'

    ## Write contents ##
    i = 2

    for result in crawl_yml_tree(yml):
        ws[PATH_COL + str(i)] = result[0]
        ws[SOURCE_COL + str(i)] = result[1]
        ws[TRANSLATED_COL + str(i)] = ''
        i += 1

    return wb

def insert_index_value_into_list(index, value, data):
    if len(data) < index+1:
        data.extend([''] * (index + 1 - len(data)))

    data[index] = value
    return data

def insert_path_value_into_dict(path, value, data):
    path_parts = path.split('.')
    key = path_parts[0]

    if len(path_parts) == 1:
        # Handle list entry:
        if "[" in key: 
            list_index = int(key[key.find('[')+1:key.find(']')])
            key = key[:key.find('[')]
            data[key] = insert_index_value_into_list(list_index, value, data[key] if key in data else list())
        else: 
            data[key] = value
    elif len(path_parts) > 1:
        data[key] = insert_path_value_into_dict(".".join(path_parts[1:]), value, data[key] if key in data else dict())
    else:
        raise Exception('ERROR: Invalid path.')

    return data

    
def convert_xlsx_to_yml(xlsx):
    ws = xlsx.active

    data = {}

    for row in ws.iter_rows(2):
        path = row[0].value
        #untranslated = row[1].value
        translated = row[2].value

        # Skip untranslated.
        if XLSX_SKIP_UNTRANSLATED and translated is None:
            continue

        data = insert_path_value_into_dict(path, translated, data)

    return data

### Options ###
PATH_COL = 'A'
SOURCE_COL = 'B'
TRANSLATED_COL = 'C'
XLSX_SKIP_UNTRANSLATED = True


if __name__ == '__main__':
    # Validate arguments
    if len(sys.argv) != 3:
        print('ERROR: Wrong number of arguments.\n')
        print_usage()
        exit(1)
    
    # Get arguments.
    input_file = sys.argv[1]
    output_file = sys.argv[2]

    input_extension = input_file.split('.')[-1]

    ### Processing ###
    # NOTE: Operation type is determined by the extension of the input file.

    if input_extension.lower() == 'yml': # YML to XLSX
        with open(input_file, 'r') as fin:
            input_data = yaml.load(fin)

        workbook = convert_yml_to_xlsx(input_data)
        
        output_file = force_extension(output_file, 'xlsx')
        workbook.save(output_file)
    elif input_extension.lower() == 'xlsx': # XLSX to YML
        input_data = load_workbook(filename = input_file)

        yml = convert_xlsx_to_yml(input_data)
    
        output_file = force_extension(output_file, 'yml')
        with open(output_file, 'w') as fout:
            yaml.dump(yml, fout, default_flow_style=False, explicit_start=True)
    else:
        print('ERROR: Unknown input file extension.')
        print_usage()
        exit(2)
